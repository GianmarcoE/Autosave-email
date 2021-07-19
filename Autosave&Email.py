import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
import os
import tkinter as tk
from tkinter import filedialog
from tkinter import *
from tkinter.ttk import *
from datetime import datetime, date
import time
from smtplib import SMTP_SSL as SMTP
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

try:
    from ctypes import windll
    windll.shcore.SetProcessDpiAwareness(1)
except:
    pass

root= tk.Tk()

root.title('Autosave&Mail')

root.geometry('550x550')
root.minsize(550, 550)
root.configure(background='#fff')

def getFile ():

    try:

        import_file_path = filedialog.askopenfilename() #Opens file path for report
        
        wb = openpyxl.load_workbook(import_file_path) #Imports the chosen file

        my_app = openpyxl.load_workbook('My_app.xlsx') #imports personal data from Excel (My_app)

        sheet = my_app["Pers Data for App (signature)"] #selects sheet from personal Excel data

        my_name = sheet.cell(row=2, column=1).value #stores name from Excel

        my_mail = sheet.cell(row=2, column=2).value  #stores email from Excel

        my_jobtitle = sheet.cell(row=2, column=3).value  #stores job title from Excel

        my_address = sheet.cell(row=2, column=4).value  #stores address from Excel

        my_password = sheet.cell(row=2, column=5).value  #stores password from Excel

        my_mnumber = sheet.cell(row=2, column=6).value #stores M number from Excel
        
        labelinfo = wb.create_sheet()  #created new Excel sheet
        labelinfo.title = "Info" #names new Excel sheet

        labelinfo['B3'] = f'Prepared by {my_name}' #writes in the given Excel cells

        labelinfo['B4'] = 'Analytics & Reporting Team, Group People' #writes in the given Excel cells

        labelinfo['B6'] = 'If you have any further questions, please contact me directly:' #writes in the given Excel cells
        labelinfo['B6'].font = Font(italic=True)

        labelinfo['B7'] = my_mail #writes in the given Excel cells
        labelinfo['B7'].font = Font(underline='single')

        for i in range (3, 7):
            labelinfo.cell(row=3, column=i).border = Border(top=Side(style='thin')) #applies borders to the Excel info sheet
        for i in range (3, 7):
            labelinfo.cell(row=7, column=i).border = Border(bottom=Side(style='thin')) #applies borders to the Excel info sheet
        for i in range (4, 7):
            labelinfo.cell(row=i, column=2).border = Border(left=Side(style='thin')) #applies borders to the Excel info sheet
        for i in range (4, 7):
            labelinfo.cell(row=i, column=7).border = Border(right=Side(style='thin')) #applies borders to the Excel info sheet

        labelinfo.cell(row=3, column=2).border = Border(top=Side(style='thin'), left=Side(style='thin')) #applies borders to the Excel info sheet
        labelinfo.cell(row=3, column=7).border = Border(top=Side(style='thin'), right=Side(style='thin')) #applies borders to the Excel info sheet
        labelinfo.cell(row=7, column=2).border = Border(bottom=Side(style='thin'), left=Side(style='thin')) #applies borders to the Excel info sheet
        labelinfo.cell(row=7, column=7).border = Border(bottom=Side(style='thin'), right=Side(style='thin')) #applies borders to the Excel info sheet
        labelinfo.sheet_view.showGridLines = False #hides gridlines from Excel sheet

        today = date.today() #imports today's date in the below formats
        d1 = today.strftime("%m.%Y")
        d2 = today.strftime("%Y.%m.%d")
        d3 = today.strftime("%Y")
            
        labelinfo['B10'] = (f'Structure: {structure.get()}') #writes user input in the given cell
        labelinfo['B10'].font = Font(bold=True)
        labelinfo['B11'] = (f'Date as of: {d2}') #writes today's date in the given cell
        labelinfo['B11'].font = Font(bold=True)
        labelinfo['B12'] = (f'Datasource: {sourcein.get()}') #writes user input in the given cell
        labelinfo['B12'].font = Font(bold=True)

        labelinfo.protection.sheet = True
        labelinfo.protection.enable() #activates Excel's sheet protection
        labelinfo.protection.password = my_password #sets password
        
        if location.get() == 'PL' and var.get() == 1: #saves file in the correct folder
            os.chdir(f"K:\\2.6 NOC HRO\\11. Processes\\Reporting\\2. Sent reports\\PL\\{d3}\\{d1}\\DLP")
            wb.save(os.path.basename(import_file_path))
        elif location.get() == 'PL' and var.get() == 0:
            os.chdir(f"K:\\2.6 NOC HRO\\11. Processes\\Reporting\\2. Sent reports\\PL\\{d3}\\{d1}")
            wb.save(os.path.basename(import_file_path))
        elif location.get() == 'SE' and var.get() == 1:
            os.chdir(f"K:\\2.6 NOC HRO\\11. Processes\\Reporting\\2. Sent reports\\SE\\{d3}\\{d1}\\DLP")
            wb.save(os.path.basename(import_file_path))
        elif location.get() == 'SE' and var.get() == 0:
            os.chdir(f"K:\\2.6 NOC HRO\\11. Processes\\Reporting\\2. Sent reports\\SE\\{d3}\\{d1}")
            wb.save(os.path.basename(import_file_path))
        elif location.get() == 'DK' and var.get() == 1:
            os.chdir(f"K:\\2.6 NOC HRO\\11. Processes\\Reporting\\2. Sent reports\\DK\\{d3}\\{d1}\\DLP")
            wb.save(os.path.basename(import_file_path))
        elif location.get() == 'DK' and var.get() == 0:
            os.chdir(f"K:\\2.6 NOC HRO\\11. Processes\\Reporting\\2. Sent reports\\DK\\{d3}\\{d1}")
            wb.save(os.path.basename(import_file_path))
        elif location.get() == 'NO' and var.get() == 1:
            os.chdir(f"K:\\2.6 NOC HRO\\11. Processes\\Reporting\\2. Sent reports\\NO\\{d3}\\{d1}\\DLP")
            wb.save(os.path.basename(import_file_path))
        elif location.get() == 'NO' and var.get() == 0:
            os.chdir(f"K:\\2.6 NOC HRO\\11. Processes\\Reporting\\2. Sent reports\\NO\\{d3}\\{d1}")
            wb.save(os.path.basename(import_file_path))
        elif location.get() == 'FI' and var.get() == 1:
            os.chdir(f"K:\\2.6 NOC HRO\\11. Processes\\Reporting\\2. Sent reports\\FI\\{d3}\\{d1}\\DLP")
            wb.save(os.path.basename(import_file_path))
        elif location.get() == 'FI' and var.get() == 0:
            os.chdir(f"K:\\2.6 NOC HRO\\11. Processes\\Reporting\\2. Sent reports\\FI\\{d3}\\{d1}")
            wb.save(os.path.basename(import_file_path))
        elif location.get() == 'EE' and var.get() == 1:
            os.chdir(f"K:\\2.6 NOC HRO\\11. Processes\\Reporting\\2. Sent reports\\EE\\{d3}\\{d1}\\DLP")
            wb.save(os.path.basename(import_file_path))
        elif location.get() == 'EE' and var.get() == 0:
            os.chdir(f"K:\\2.6 NOC HRO\\11. Processes\\Reporting\\2. Sent reports\\EE\\{d3}\\{d1}")
            wb.save(os.path.basename(import_file_path))
        elif location.get() == 'Learning Reports' and var.get() == 1:
            os.chdir(f"K:\\2.6 NOC HRO\\11. Processes\\Reporting\\2. Sent reports\\Learning Reports\\{d3}\\{d1}\\DLP")
            wb.save(os.path.basename(import_file_path))
        elif location.get() == 'Learning Reports' and var.get() == 0:
            os.chdir(f"K:\\2.6 NOC HRO\\11. Processes\\Reporting\\2. Sent reports\\Learning Reports\\{d3}\\{d1}")
            wb.save(os.path.basename(import_file_path))
        elif location.get() == 'Other' and var.get() == 1:
            os.chdir(f"K:\\2.6 NOC HRO\\11. Processes\\Reporting\\2. Sent reports\\Other\\{d3}\\{d1}\\DLP")
            wb.save(os.path.basename(import_file_path))
        elif location.get() == 'Other' and var.get() == 0:
            os.chdir(f"K:\\2.6 NOC HRO\\11. Processes\\Reporting\\2. Sent reports\\Other\\{d3}\\{d1}")
            wb.save(os.path.basename(import_file_path))
        elif location.get() == 'TA Report':
            os.chdir("K:\\2.6 NOC HRO\\11. Processes\\Reporting\\TA Reporting\\TA Reporting - sent reports")
            wb.save(os.path.basename(import_file_path))

        def smtp_endpoint(): #estabilishes SMTP connection
            smtp = os.getenv('EMAIL_NOTIFICATIONS_SMTP_ENDPOINT', 'ccdfooo:25') #if the mails don't work, the IP is shut down, we should ask for a new one.
            return tuple(smtp.split(':'))

        def sender(): #defines email sender
            return str("peoplereports@foo.com")
            #return getpass.getuser() + '@' + socket.gethostname()

        def send_email(send_to, subject, message, file_path=None, file_name=None): #function to send email
            send_from = sender()
            msg = MIMEMultipart('alternative')
            msg['From'] = send_from
            msg['To'] = send_to + '; peoplereports@foo.com' #send email also to our shared mailbox - further rules are applied on Outlook so the mail moves automatically to the sent folder
            msg["Cc"] = cc.get()
            msg["Bcc"] = bcc.get()
            msg['Subject'] = subject
            msg.attach(MIMEText(message, 'html'))
            if file_path:
                f = open(file_path, 'rb')
                attachment = MIMEApplication(f.read())
                #attachment = MIMEText(f.read())
                email_filename = file_name if file_name else file_path
                attachment.add_header('Content-Disposition', 'attachment', filename=email_filename)
                msg.attach(attachment)
            smtp = smtp_endpoint()
            server = smtplib.SMTP(smtp[0], smtp[1])
            server.sendmail(send_from, msg['To'].split(";") + msg["Cc"].split(";") + msg["Bcc"].split(";"), msg.as_string()) #actually sends email to receiver/Cc/Bcc separated by a ";"
            server.close()

        time.sleep(3) #HTML & CSS Email text
        message = f"""\
        <html>
        Hi,
        <br><br>
        Please find attached the requested report.
        <br><br>
        <div style="font-size:90%; color:#0606bf;">
            <b>NEED A REPORT?</b>
        </div>
        <div style="font-size:90%;">
            <i>I really enjoy helping our people â to be sure that you get your report as soon as possible, please use THIS tool instead of sending request via e-mail.</i>
        </div>
        <br>
        Kind Regards,
        <br><br>
        <b>{my_name}</b>
        <br>
        <div style="font-size:90%;">
            {my_jobtitle}
        <div>
        <br>
        <div style="font-size:90%;">
            <b style="color:#0606bf;">Foo</b> | Analytics & Reporting
        </div>
        </div>
        <div style="font-size:90%;">
            Visit me: {my_address}
        </div>
        <div style="font-size:90%;">
            E-mail: {my_mail}
        </div>
        <div style="font-size:90%;">
            Web:
            <a href="www.foo.com" target="_blank">
                foo.com
            </a>
        </div>
        <br>
        <div style="font-size:85%; color:#888">
            Foo info
            <br><br>
            This e-mail may contain confidential information. If you receive this e-mail by mistake, please inform the sender, delete the e-mail and do not share or copy it.
        </div>
        </html>
        """
        send_email(email.get(), f"AutoPY {os.path.basename(import_file_path)}", message, f"{os.path.abspath(os.path.basename(import_file_path))}", f"{os.path.basename(import_file_path)}")

        os.chdir(f"C:\\Users\\{my_mnumber}\\Desktop") #set the directory back to desktop, so we'll be able to find the My_app Excel file

        email.delete(0, 'end') #cancels what was written in the email address field
        success = 'Report succesfully sent!'
        email.insert(0, success) #inserts in the email address field the success message

    except: #if any error occurs
        email.delete(0, 'end') #cancels what was written in the email address field
        error = 'An error occured, please try again'
        email.insert(0, error) #inserts in the email address field the error message
        
structure = Combobox(root, width=8) #values for "structure" dropdown menu
structure ['values']= ("MCC", "Org")
structure.configure(font=10)
structure.current(0)

location = Combobox(root, width=8) # values for "folder" dropdown menu
location ['values']= ("PL", "SE", "DK", "NO", "FI", "EE", "Other", "Learning Reports", "TA Report")
location["state"]= "readonly"
location.configure(font=10)
location.current(0)

infolbl = Label(root, text="Signature", font=('helvetica 8 bold'), foreground='#008000', background='#fff')
infolbl.grid(row=0, sticky='W', padx=(10, 0))
Separator(root, orient="horizontal")
Separator().grid(row=0, column= 0, pady= (20, 0), sticky=EW, columnspan=2)

structurelbl = Label(root, text = "Structure", font=('helvetica', 10), background='#fff')
structurelbl.grid(row=2, sticky='E', padx= 15, pady=(20, 0))

source = Label(root, text = "Datasource", font=('helvetica', 10), background='#fff')
source.grid(row= 3, sticky='E', padx= 15, pady=20)

locationlbl = Label(root, text = "Folder", font=('helvetica', 10), background='#fff')
locationlbl.grid(row=5, column= 0, sticky='E', padx= 15, pady=20)

infolbl = Label(root, text="Save to", font=('helvetica 8 bold'), foreground='#008000', background='#fff')
infolbl.grid(row=4, sticky='W', padx=(10, 0))
Separator(root, orient="horizontal")
Separator().grid(row=4, column= 0, pady=(20, 0), sticky=EW, columnspan=2)

structure.grid(row= 2, column= 1, padx=15, pady=(20, 0))

sourcein = Entry(root, width=25, font=('helvetica', 10))
sourcein.grid(row= 3, column= 1, padx=(15, 50), pady=20, sticky='E')

location.grid(row= 5, column= 1, padx=15, pady=20)

var = IntVar()
dlp = tk.Checkbutton (root, text='DLP', variable=var, bg='#fff', font=('helvetica', 10)) #tickbox for DLP option in "Folder"
dlp.grid(row=5, column=1, padx=(0, 50), pady=20, sticky='E')

infolbl = Label(root, text="Send Email", font=('helvetica 8 bold'), foreground='#008000', background='#fff')
infolbl.grid(row=7, sticky='W', padx=(10, 0))
Separator(root, orient="horizontal")
Separator().grid(row=7, column= 0, pady= (20, 0), sticky=EW, columnspan=2)

emailto = Label(root, text = "To", font=('helvetica', 10), background='#fff')
emailto.grid(row=8, column=0, sticky='E', padx= 15, pady=(20, 0))
email = Entry(root, width=25, font=('helvetica', 10))
email.grid(row=8, column= 1, padx=(15, 50), pady=(20, 0), sticky='E')

ccto = Label(root, text = "Cc", font=('helvetica', 10), background='#fff')
ccto.grid(row=9, column=0, sticky='E', padx= 15, pady=(20, 0))
cc = Entry(root, width=25, font=('helvetica', 10))
cc.grid(row=9, column= 1, padx=(15, 50), pady=(20, 0), sticky='E')

bccto = Label(root, text = "Bcc", font=('helvetica', 10), background='#fff')
bccto.grid(row=10, column=0, sticky='E', padx= 15, pady=(20, 0))
bcc = Entry(root, width=25, font=('helvetica', 10))
bcc.grid(row=10, column= 1, padx=(15, 50), pady=(20, 0), sticky='E')

browseButton_Excel = tk.Button(text='Choose File', command=getFile, bg='#008000', fg='#fff', font=('helvetica', 12))
browseButton_Excel.grid(columnspan=2, pady= 30)

root.columnconfigure(0, weight=1)
root.columnconfigure(1, weight=2)
root.rowconfigure(11, weight=1)

root.mainloop()
